from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

pergunta1 = input("qual o nome do arquivo com as informações? (insira o nome sem extensão)\n> ")
txt_path = f"{pergunta1}.txt"
pergunta2 = input("como quer salvar sua planilha? (insira o nome sem extensão)\n> ")
xlsx_out = f"{pergunta2}.xlsx"

def split_csv_line(line: str) -> list[str]:
    """
    Divide por vírgula (CSV simples). Remove espaços nas bordas.
    Se você precisar suportar vírgulas dentro de aspas, eu adapto usando o módulo csv.
    """
    return [p.strip() for p in line.strip().split(",")]

with open(txt_path, "r", encoding="utf-8") as f:
    lines = [ln.rstrip("\n") for ln in f]

if len(lines) < 3:
    raise ValueError("O TXT precisa ter pelo menos 3 linhas: cabeçalho (linha 1), linha 2 ignorada, e dados a partir da linha 3.")

# 1) Cabeçalho = primeira linha
headers = split_csv_line(lines[0])
if not headers or any(h == "" for h in headers):
    raise ValueError("Cabeçalho inválido na 1ª linha. Verifique se há colunas separadas por vírgula.")

# 2) Dados = terceira linha em diante
rows = []
for idx, line in enumerate(lines[2:], start=3):
    if not line.strip():
        continue # pular linhas vazias
    values = split_csv_line(line)

    if len(values) != len(headers):
        raise ValueError(
            f"Linha {idx} tem {len(values)} colunas, mas o cabeçalho tem {len(headers)}.\n"
            f"Linha problemática: {line}"
        )

    rows.append(tuple(values))

# 3) Criar Excel
wb = Workbook()
ws = wb.active
ws.title = "Ordem do Culto"

ws.append(headers)

# Estilo do cabeçalho
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Inserir linhas
for r in rows:
    ws.append(r)

# Ajustar largura de colunas automaticamente
for col in range(1, len(headers) + 1):
    max_len = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col, max_col=col):
        val = row[0].value
        if val is None:
            continue
        max_len = max(max_len, len(str(val)))
    ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60) # limite 60

wb.save(xlsx_out)
print("Excel gerado com sucesso:", xlsx_out)
print("Total de registros importados:", len(rows))
