from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Criar planilha
wb = Workbook()
ws = wb.active
ws.title = "Ordem Evento"

headers = []

while True:
    print("-----------------------------------------------------\n")
    print("Adicione o cabeçalho da planilha.")
    print("\n-----------------------------------------------------")
    
    item = input("Coloque um item para o Cabeçalho como:\n(Horário, Evento, Área Responsável):\n> ")
    
    if not item:
        print("Cabeçalho vazio não é permitido.")
        continue
    
    headers.append(item)
    
    pergunta = input("Deseja adicionar outro item? (s/n): ").strip().lower()
    if pergunta == 's':
        continue
    else:
        break
    
ws.append(headers)

# Estilo do cabeçalho
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Congelar linha 1 (cabeçalho)
ws.freeze_panes = "A2"

# Ativar filtro no cabeçalho
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

print("\n-----------------------------------------------------")
print(" Agora, insira as atividades de cada item colocado anteriormente!")
print("-----------------------------------------------------")

while True:
    row = []
    print("\n------------- Nova atividade/linha -------------\n")

    for h in headers:
        valor = input(f"{h}: ").strip()
        row.append(valor)

    ws.append(row)

    pergunta = input("Deseja adicionar outra atividade? (s/n): ").strip().lower()
    if pergunta == 's':
        continue
    else:
        break

# Ajustar largura de colunas automaticamente
for col in range(1, len(headers) + 1):
    col_letter = get_column_letter(col)

    max_len = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        v = str(v)
        if len(v) > max_len:
            max_len = len(v)

    ws.column_dimensions[col_letter].width = min(max_len + 5, 60)  # limite 60

print()
nome = input("Como deseja que a planilha seja salva?\nExemplo (Ordem_Eventos)\n> ").strip().lower()
file_path = f"{nome}.xlsx"

wb.save(file_path)

print(f"\n Planilha salva como {file_path}")