from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Criar planilha
wb = Workbook()
ws = wb.active
ws.title = "Ordem Evento"

# Cabeçalhos
headers = ["Horário", "Evento", "Área Responsável"]
ws.append(headers)

# Estilo do cabeçalho
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

# Dados
rows = [
]

while True:
    
    print("-----------------------------------------------------\n")
    print("Adicione uma nova atividade.")
    print("\n-----------------------------------------------------")
    
    horario_inicio = input("Horário inicio: ")
    horario_final = input("Horário final: ")
    
    print("\n-----------------------------------------------------\n")
    evento = input("Evento: ")
    departamento = input("Área Responsável: ")
    print("\n-----------------------------------------------------\n")
    
    rows.append((f"{horario_inicio} - {horario_final}", evento, departamento))
    
    pergunta = input("Deseja adicionar outra atividade? (s/n): ").strip().lower()
    
    if pergunta == 's':
        continue
    else:
        break

for x in rows:
    ws.append(x)

# Ajustar largura das colunas
for col in range(1, 4):
    ws.column_dimensions[get_column_letter(col)].width = 35

file_path = "Ordem_Evento.xlsx"
wb.save(file_path)

print(f"Planilha salva como {file_path}")

file_path